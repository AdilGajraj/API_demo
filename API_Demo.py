import requests
import json
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, Border, Side

# assign url to variable
import requests

url = "https://motorcycle-specs-database.p.rapidapi.com/category"

headers = {
	"X-RapidAPI-Host": "motorcycle-specs-database.p.rapidapi.com",
	"X-RapidAPI-Key": "c6d3bf4ed2mshdabd204c75b404dp15dc0bjsn61acd9c567ed"
}

response = requests.request("GET", url, headers=headers)
# Check response status of url
# print response
print(response)
# print content / response.text
print(response.text)
# assign response object ariable (json loads)
clean_data = json.loads(response.text)
# print the clean data
print(clean_data)

for i in range(len(clean_data)):
    result = clean_data[i]['name']
    print(result)
# create a workbook
wb = Workbook()
ws = wb.active
sheet = wb['Sheet1']
# create a spreadsheet page

sheet["A1"] = 'ID'
sheet['B1'] = 'Name'

# designate columns
                 
header_list = [sheet['A1'], sheet['B1']]
for header in header_list:
    header.font = Font(bold = True)
# Loop through the columns
    for i in range(len(header_list)):
        id = clean_data[i]['id']
        name = clean_data[i]['name']
        # Assign the data to the spreadsheet
        sheet.cell(row = i+2, column = 1).value = id
        sheet.cell(row = i+2, column = 2).value = name
# fill the cell values with appropriate data

# Save the workbook
wb.save('API Demo.xlsx')